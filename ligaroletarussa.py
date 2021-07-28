from sqlite3.dbapi2 import IntegrityError, OperationalError
import cartolafc
import sqlite3
from openpyxl.descriptors import base
import requests
from os import system
from openpyxl import load_workbook
from random import randint

opcoes_de_ligas = [{'opção':'Liga Principal'},
                   {'opção':'Liga Eliminatória'}]


class Exibir:
    """
    Funções para exibir algo na tela.
    """
    @staticmethod
    def exibir_cabecalho(texto:str =''):
        system('cls') or None
        print('-'*50)
        print(f"|{texto.upper():^48}|")
        print('-'*50)

    @staticmethod
    def listar_itens(lista:list):
        """
        Método que exibe na tela uma lista tabelada das opções.
        
        Recebe:     lista - uma lista de dicionários contendo as opções.
                    As chaves serão os cabeçalhos da tabela.
        """
        chaves = [key for key in lista[0].keys()]
        maiores = []
        for chave in chaves:
            valor_maior = 0
            for item in lista:
                if len(str(item[chave])) > valor_maior:
                    valor_maior = len(str(item[chave]))
                else:
                    pass
            maiores.append(valor_maior)
        print(f" {'No':_^4} ", end="")
        for c in range(0, len(chaves)):
            print(f" {chaves[c].upper():_^{maiores[c]+2}} ", end="")
        print()
        numero = 1
        for dic in lista:
            print(f" {(numero):^4} ", end="")
            for c in range(0, len(chaves)):
                print(f" {dic[chaves[c]]:<{maiores[c]+2}} ", end="")
            print()
            numero += 1
        print(f" {'0':^4}  Nenhuma das alternativas")
   
    @staticmethod
    def int_input(texto:str =''):
        """
        Método para garantir que a entada será um inteiro.
        """
        while True:
            try:
                entrada = int(input(texto))
            except ValueError:
                print('Digite um valor válido.')
            else:
                break
        return entrada

    def escolher_entre_opcoes(self, dicionario:list):
        """"
        Método que devolve a opção escolhida.
        
        Recebe:     dicionario - lista de dicionários com as opções.
        
        Retorna:    dicionário com a opção escolhida ou um dicionário vazio.
        """
        while True:
            escolha = self.int_input(f'Escolha uma das opções entre 1 e {len(dicionario)} ou 0 e digite aqui sua escolha: ')
            if escolha == 0:
                return {}
            elif 0 < escolha <= len(dicionario):
                return dicionario[escolha-1]
            else:
                print('Opção inválida.')
    
    def escolher_ligas_roleta_russa(self):
        """
        Método que lista e rece a escolha entre as ligas da Roleta Ru$$a.
        
        Retorna:    string com o nome da tabela escolhida ou uma string vazia.
        """
        self.listar_itens(opcoes_de_ligas)
        escolha = self.escolher_entre_opcoes(opcoes_de_ligas)
        if not escolha:
            return ''
        else:
            return escolha['opção'].replace(' ', '').replace('ó', 'o')


class RoletaRussa:
    """Classe principal, para servir de base para as outras."""
    def __init__(self):
        self.banco_de_dados = "ligaroletarussa2021.db"

    def acessar_banco_de_dados(self):
        """
        Acessa o Banco de Dados.
        
        Retorna:    con - Conecção do Banco de Dados.
                    cursor - Método para utilizar instruções do Banco de Dados.
        """
        con = sqlite3.connect(self.banco_de_dados)
        cursor = con.cursor()
        return con, cursor

    def pegar_autenticacao(self):
        """
        Acessa o Banco de Dados e pega o 'código de autorização'.
        
        Retorna:    auth - string com o código.
        """
        con, cursor = self.acessar_banco_de_dados()
        cursor.execute("SELECT cookie FROM Auth")
        auth = cursor.fetchall()
        return auth[0][0]
    
    def trocar_autenticacao(self):
        """
        Caso o código de autenticação não funcione, este método auxilia na troca do código.
        
        Para pegar o código, acesse o site do Cartola FC pelo navegador e faça o login.
        Após o login, acesse a página https://login.globo.com/api/user e copie o código apresentado
        na variável glbId.
        """
        nova_autenticacao = str(input('Digite a nova autenticação: '))
        con, cursor = self.acessar_banco_de_dados()
        cursor.execute(f'UPDATE Auth SET cookie={nova_autenticacao}')
        con.commit()
    
    def acesso_autenticado(self):
        """
        Acessa o API do Cartola FC.
        
        Retorna:    api - Instancia da classe API da biblioteca cartolafc.
        """
        api = cartolafc.Api()
        while True:
            try:
                api._glb_id = self.pegar_autenticacao()
                api.ligas(query='')
            except:
                print("Erro de login.")
                self.trocar_autenticacao()
            else:
                break
        return api
    
    def pesquisar_time(self, termo_pesquisa:str =None):
        """
        Pesquisa por um time na API do CArtola FC.
        
        Recebe:     termo_pesquisa(opcional) - termo que será utilizado para efetuar a pesquisa.
        
        Retorna:    lista_times - uma lista de dicionários contendo os dados dos times escontrados.
        """
        lista_times = []
        api = self.acesso_autenticado()
        if not termo_pesquisa:
            termo_pesquisa = str(input('Digite o nome do Time: ')).strip()
        times = api.times(query=termo_pesquisa)
        for item in times:
            temp = {"id":item.id, "nome":item.nome, "cartoleiro":item.nome_cartola}
            lista_times.append(temp)
        return lista_times


class CadastroTime(RoletaRussa):
    """
    Cadastrar um time no Banco de Dados.
    Basta instanciar a classe que o processo de cadastro se inicia automaticamente.
    """
    def __init__(self, dados:dict =None, tabela:str =None):
        super().__init__()
        tela = Exibir()
        if not dados:
            while True:
                lista_times = self.pesquisar_time()
                tela.listar_itens(lista_times)
                time = tela.escolher_entre_opcoes(lista_times)
                if time:
                    break
        else:
            time = dados
        if not tabela:
            while True:
                tabela = tela.escolher_ligas_roleta_russa()
                if tabela:
                    break
        else:
            tabela = tabela
        self.cadastrar_time_no_BD(tabela=tabela, dados=time)
    
    def cadastrar_time_no_BD(self, tabela:str, dados:dict):
        """
        Guarda os dados de um time no Banco de Dados.
        OBS: Este método automaticamente também guarda os dados do time na tabela TimesCadastrados.
        
        Recebe:     tabela - nome da tabela onde os dadoss erão salvos.
                    dados - dicionário com os dados do time.
        
        Retorna:    mensagem de sucesso ou falha.
        """
        dados_time = (dados['id'], dados['nome'], dados['cartoleiro'])
        con, cursor = self.acessar_banco_de_dados()
        try:
            cursor.execute(f'INSERT INTO {tabela}(ID, Nome, Cartoleiro) VALUES{dados_time}')
        except IntegrityError:
            print(f'Já existe um cadastro do time {dados_time[1]}.')
            pass
        except OperationalError:
            print(f'Não existe a tabela {tabela}.')
            pass
        else:
            con.commit()
        if tabela in ('LigaPrincipal', 'LigaEliminatória'):
            dados_time2 = (dados['id'], dados['nome'], dados['cartoleiro'], 'OK')
            try:
                cursor.execute(f'INSERT INTO TimesCadastrados(ID, Nome, Cartoleiro, {tabela}) VALUES{dados_time2}')
            except IntegrityError:
                cursor.execute(f'UPDATE TimesCadastrados SET {tabela}="OK" WHERE ID={dados_time2[0]}')
                con.commit()
            except OperationalError:
                print(f'Não existe a coluna {tabela} em TimesCadastrados.')
                pass
            else:
                con.commit()
        print(f'Time {dados_time[1]} foi cadastrado com sucesso.')


class CadastroTimesLiga(RoletaRussa):
    """
    Cadastra todos os times de uma liga..
    Basta instanciar a classe e o processo de cadastro começa.
    """
    def __init__(self):
        super().__init__()
        tela = Exibir()
        while True:
            lista_ligas = self.pesquisar_liga()
            tela.listar_itens(lista_ligas)
            liga = tela.escolher_entre_opcoes(lista_ligas)
            if liga:
                break
        while True:
            tabela = tela.escolher_ligas_roleta_russa()
            if tabela:
                break
        self.cadastrar_times_de_liga_BD(liga=liga, tabela=tabela)
        if tabela == 'LigaPrincipal':
            self.cadastrar_times_de_liga_BD(liga=liga, tabela='Patrimonio')
            self.cadastrar_times_de_liga_BD(liga=liga, tabela='Turno')
            self.cadastrar_times_de_liga_BD(liga=liga, tabela='Returno')


    def pesquisar_liga(self, termo_pesquisa:str =None):
        """
        Pesquisa por uma liga na API do CArtola FC.
        
        Recebe:     termo_pesquisa(opcional) - termo que será utilizado para efetuar a pesquisa.
        
        Retorna:    lista_ligas - uma lista de dicionários contendo os dados das lligas encontradas.
        """
        lista_ligas = []
        api = self.acesso_autenticado()
        if not termo_pesquisa:
            termo_pesquisa = str(input('Digite o nome da Liga: ')).strip()
        ligas = api.ligas(query=termo_pesquisa)
        for item in ligas:
            temp = {"nome":item.nome, "slug":item.slug}
            lista_ligas.append(temp)
        return lista_ligas
    
    def pegar_times_liga(self, liga:dict):
        """
        Método pra pegar todos os times da liga.
        
        Recebe:     liga - dicionário com os dados da liga
        
        Retorna:    times_liga.times - uma lista de dicionários contendo os dados dos times.
        """
        api = self.acesso_autenticado()
        liga_slug = str(liga['slug'])
        times_liga = api.liga(slug=liga_slug)
        return times_liga.times

    def cadastrar_times_de_liga_BD(self, liga:dict, tabela:str):
        """
        Cadastra os times da liga no Banco de Dados.
        
        Recebe:     liga - dicionário com os dados da liga.
                    tabela - string com o nome da tabela onde os dados serão salvos.
        
        Retorna:    mensagem de sucesso
        """
        times = self.pegar_times_liga(liga)
        for time in times:
            t = {'id':time.id, 'nome':time.nome, 'cartoleiro':time.nome_cartola}
            CadastroTime(dados=t, tabela=tabela)
        print('Times da liga foram salvos.')


class Pontuacao(RoletaRussa):
    """
    Classe base para o processo de coleta de pontuação dos times.
    """
    def __init__(self):
        super().__init__()
        self.rodada_atual = self.rodada() - 1

    def pegar_pontuacao_dos_times(self, tabela:str):
        """
        Acessa a tabela do Bando de Dados, pega os times cadastrados e busca suas pontuações da API.
        
        Recebe:     tabela - string com o nome da tabela do Banco de Dados.
        
        Retorna:    pontuacoes - lista de dicionários contendo id do time, pontuação da rodada e
                                patrimônio.
        """
        con, cursor = self.acessar_banco_de_dados()
        cursor.execute(f'select ID from {tabela}')
        lista_id = cursor.fetchall()
        api = self.acesso_autenticado()
        pontuacoes = []
        for id in lista_id:
            try:
                t = api.time(id=id[0], as_json=True)
            except cartolafc.CartolaFCError:
                t = {'pontos':0, 'patrimonio':0}
            pontuacoes.append({'id':id[0], 'pontos':t['pontos'], 'patrimonio':t['patrimonio']})
        return pontuacoes

    @staticmethod
    def rodada(): 
        """
        Pega o número da rodada atual.
        
        Retorna:    status['rodada_atual'] - inteiro representando a rodada atual.
        """
        try:
            url_1 = 'https://api.cartolafc.globo.com/mercado/status'
            r = requests.get(url=url_1)
            status = r.json()
        except:
            print('Não foi possível pegar a rodada no sistema.')
            tela = Exibir()
            return tela.int_input('Digite a rodada:')
        else:
            return status['rodada_atual']

    def salvar_pontos(self, tabela:str, coluna:str, pontos:dict):
        """
        Salva a pontuação do time no Banco de Dados.
        
        Recebe:     tabela - string com o nome da tabela onde os dados serão salvos.
                    coluna - string com a coluna onde os dados serão salvos.
                    pontos - dicionário com os dados do time que serão salvos.
        """
        con, cursor = self.acessar_banco_de_dados()
        try:
            cursor.execute(f"UPDATE {tabela} SET {coluna}={pontos['pontos']} WHERE ID={pontos['id']}")
        except OperationalError:
            print(f'Coluna {coluna} da talela {tabela} não existe.')
            pass
        else:
            con.commit()

    def numero_de_eliminados(self):
        """
        Define quantos times serão eliminados da Liga Eliminatória de acordo com a Rodada.

        Retorna:    numero_eliminados - inteiro com o número de eliminados
        """
        if 6 <= self.rodada_atual <= 24:
            numero_eliminados = 1
        elif 25 <= self.rodada_atual <= 37:
            numero_eliminados = 2
        else:
            numero_eliminados = 0
        return numero_eliminados


class PontosLigaPrincipal(Pontuacao):
    """
    Salva as Pontuações da Liga Principal.
    Basta instanciar que o processo começa automaticamente.
    """
    def __init__(self):
        super().__init__()
        self.coluna = f'Rodada{self.rodada_atual}'
        pontos = self.pegar_pontuacao_dos_times(tabela='LigaPrincipal')
        self.salvar_principal(pontos=pontos)
        self.salvar_mito()
        print('Pontuação dos times da Liga Principal salvas com SUCESSO.')
    
    def salvar_principal(self, pontos:list):
        """
        Salva as pontuações.
        OBS: Este método salva automativamente as pontuações das tabelas Turno e Returno, na
            coluna que foi informada.
        
        Recebe:     coluna - coluna da tabela do Banco de Dados onde os dados serão salvos.
                    pontos - lista de dicionários com os dados a serem salvos.
        """
        for p in pontos:
            self.salvar_pontos(tabela='LigaPrincipal',coluna = self.coluna, pontos=p)
            self.salvar_patrimonio(pontos=p)
            self.salvar_turno_returno(pontos=p)
            self.salvar_pontos_total('LigaPrincipal')
            self.salvar_pontos_total('Turno')
            self.salvar_pontos_total('Returno')

    def salvar_patrimonio(self, pontos:dict):
        """
        Salva o patrimônio dos times da Liga Principal.
        
        Recebe:     pontos - dicionário
        """
        con, cursor = self.acessar_banco_de_dados()
        cursor.execute(f"UPDATE Patrimonio SET C$={pontos['patrimonio']} WHERE ID={pontos['id']}")
        con.commit()

    def salvar_mito(self):
        """
        Salva as 3 melhores pontuações da rodada na tabela Mito.
        """
        con, cursor = self.acessar_banco_de_dados()
        cursor.execute(f"SELECT ID, Nome, Cartoleiro, Rodada{self.rodada_atual} FROM LigaPrincipal ORDER BY Rodada{self.rodada_atual} DESC LIMIT 3")
        times = cursor.fetchall()
        for t in times:
            valores = (t[0], t[1], t[2], self.rodada_atual, t[3])
            cursor.execute(f"INSERT INTO Mito(ID, Nome, Cartoleiro, Rodada, Pontuacao) VALUES{valores}")
            con.commit()
        
    def salvar_turno_returno(self, pontos):
        """
        Identifica se está no primeiro turno(rodada 1 a 19) ou segundo turno(rodada 20 a 38) e
        salva os dados da rodada na tabela específica.
        
        Recebe:     coluna:"""
        if self.rodada_atual <= 19:
            self.salvar_pontos(tabela='Turno', coluna=self.coluna, pontos=pontos)
        else:
            self.salvar_pontos(tabela='Returno', coluna=self.coluna, pontos=pontos)
            
    def salvar_pontos_total(self, tabela:str):
        """
        Soma os pontos das rodadas já concluídas e salva a soma na coluna PtsTotal.

        Recebe:     tabela - tabela onde a soma será salva.
        """
        con, cursor = self.acessar_banco_de_dados()
        cursor.execute(f"SELECT ID FROM {tabela}")
        id = cursor.fetchall()
        cursor.execute(f"PRAGMA table_info({tabela})")
        col = cursor.fetchall()
        for t in id:
            pontos_total = 0
            for coluna in col[4:(self.rodada_atual + 4)]:
                cursor.execute(f"SELECT {coluna[1]} FROM {tabela} WHERE ID={t[0]}")
                pontos = cursor.fetchall()
                if pontos[0][0]:
                    pontos_total += pontos[0][0]
            cursor.execute(f"UPDATE {tabela} SET PtsTotal={pontos_total} WHERE ID={t[0]}")
            con.commit()


class PontosLigaEliminatoria(Pontuacao):
    """
    Salva os pontos dos times da Liga Eliminatória.
    Basta instanciar e o processo ocorre automaticamente.
    """
    def __init__(self):
        super().__init__()
        api = self.acesso_autenticado()
        tabela = 'LigaEliminatoria'
        con, cursor = self.acessar_banco_de_dados()
        cursor.execute(f'Select ID FROM {tabela}')
        ids = cursor.fetchall()
        for i in ids:
            id = i[0]
            time = api.time(id=id, as_json=True)
            pontos = 0
            for j in time['atletas']:
                pontos += j['pontos_num']
            cursor.execute(f'UPDATE {tabela} SET PtsRodada={pontos} WHERE ID={id}')
            con.commit()
        print('Pontuação dos times da Liga Eliminatoria salvas com SUCESSO.')


class Informativos(Pontuacao):
    """
    Atualiza a planilha de excel dos informativos.
    Basta instanciar a classe que o informativo é atualizado automaticamente.
    """
    def __init__(self):
        super().__init__()
        self.con, self.cursor = self.acessar_banco_de_dados()
        self.arquivo_xlsx = 'ligaroletarussa2021.xlsx'
        self.arquivo = load_workbook(self.arquivo_xlsx)
        self.top_10()
        self.turno_returno()
        self.patrimonio()
        self.mito()
        self.eliminatoria()
        self.matamata()
        self.arquivo.save(self.arquivo_xlsx)
        self.arquivo.close()
        print('Informativos Salvos com Sucesso.')

    def top_10(self):
        """
        Salva os dados do TOP10 na planilha.
        """
        planilha = self.arquivo['LigaPrincipal']
        self.cursor.execute("SELECT Nome, PtsTotal FROM LigaPrincipal order by PtsTotal DESC limit 10")
        top10 = self.cursor.fetchall()
        contador = 3
        for time in top10:
            planilha[f'B{contador}'] = time[0]
            planilha[f'C{contador}'] = time[1]
            contador += 1

    def turno_returno(self):
        """
        Salva os dados do Turno ou do Returno na planilha.
        """
        planilha = self.arquivo['LigaPrincipal']
        if self.rodada_atual <= 19:
            self.cursor.execute("SELECT Nome, PtsTotal FROM Turno order by PtsTotal DESC limit 3")
            turno = self.cursor.fetchall()
            contador = 3
            for time in turno:
                planilha[f'F{contador}'] = time[0]
                planilha[f'G{contador}'] = time[1]
                contador += 1
        else:
            self.cursor.execute("SELECT Nome, PtsTotal FROM Returno order by PtsTotal DESC limit 3")
            returno = self.cursor.fetchall()
            contador = 9
            for time in returno:
                planilha[f'F{contador}'] = time[0]
                planilha[f'G{contador}'] = time[1]
                contador += 1

    def patrimonio(self):
        """
        Salva os dados de patrimonio na planilha.
        """
        planilha = self.arquivo['LigaPrincipal']
        self.cursor.execute("SELECT Nome, C$ FROM Patrimonio order by C$ DESC limit 3")
        patrimonio = self.cursor.fetchall()
        contador = 15
        for time in patrimonio:
            planilha[f'F{contador}'] = time[0]
            planilha[f'G{contador}'] = time[1]
            contador += 1
    
    def mito(self):
        """
        Salva os dados de mito na planilha.
        """
        planilha = self.arquivo['LigaPrincipal']
        self.cursor.execute("SELECT Nome, Pontuacao, Rodada FROM Mito order by Pontuacao DESC limit 3")
        patrimonio = self.cursor.fetchall()
        contador = 21
        for time in patrimonio:
            planilha[f'F{contador}'] = time[0]
            planilha[f'G{contador}'] = time[1]
            planilha[f'H{contador}'] = time[2]
            contador += 1
    
    def eliminatoria(self):
        """
        Salva os dados ds Liga Eliminatória na planilha e 'elimina' o(s) time(s) com menor pontuação.
        """
        planilha = self.arquivo['LigaEliminatória']
        self.cursor.execute("SELECT Nome, PtsRodada, ID FROM LigaEliminatoria order by PtsRodada DESC")
        times = self.cursor.fetchall()
        contador = 3
        for time in times:
            planilha[f'B{contador}'] = time[0]
            planilha[f'C{contador}'] = time[1]
            contador += 1
        c = self.numero_de_eliminados()
        while True:
            if c == 0:
                break
            else:
                eliminado = times[-1]
                self.cursor.execute(f'DELETE FROM LigaEliminatoria WHERE id={eliminado[2]}')
                self.con.commit()
                times.pop()
            c -= 1
    
    def matamata(self):
        """
        Salva os nomes dos times da rodada na planilha MataMataLiga.
        """
        planilha = self.arquivo['MataMataLiga']
        coluna = f"Rodada{self.rodada_atual}"
        self.cursor.execute(f"SELECT Nome, {coluna} FROM LigaPrincipal ORDER BY {coluna} DESC")
        times = self.cursor.fetchall()
        contador = 1
        for t in times:
            planilha[f'A{contador}'] = t[0]
            contador += 1


class MataMataLiga(Pontuacao):
    def __init__(self):
        super().__init__()
    
    def pegar_times(self, a=False, b=False, c=False):
        con, cursor = self.acessar_banco_de_dados()
        cursor.execute(f"SELECT ID, Nome, Cartoleiro FROM LigaPrincipal ORDER BY Rodada{self.rodada_atual} DESC;")
        todos_times = cursor.fetchall()
        if a:
            lista_times = []
            for t in range(0, 32):
                time = todos_times[t]
                time_temp = (time[0], time[1], time[2])
                lista_times.append(time_temp)
            return lista_times
        elif b:
            lista_times = []
            for t in range(32, 64):
                time = todos_times[t]
                time_temp = (time[0], time[1], time[2])
                lista_times.append(time_temp)
            return lista_times
        elif c:
            lista_times = []
            for t in range(64, 96):
                time = todos_times[t]
                time_temp = (time[0], time[1], time[2])
                lista_times.append(time_temp)
            return lista_times
        else:
            return []

    def numeros_aleatorios(self):
        numeros_sorteados = []
        while True:
            numero = randint(0, 31)
            if numero not in numeros_sorteados:
                numeros_sorteados.append(numero)
            elif len(numeros_sorteados) == 32:
                break
        return numeros_sorteados

    def sorteio(self, lista_times, numeros_sorteados, planilha):
        arquivo = load_workbook('ligaroletarussa2021.xlsx')
        contador_planilha = 1
        plan = arquivo[planilha]
        for a  in range(0, 16):
            time = lista_times[numeros_sorteados[a]]
            plan[f'A{contador_planilha}'] = time[0]
            plan[f'B{contador_planilha}'] = time[1]
            plan[f'B{contador_planilha + 2}'] = time[2]
            contador_planilha += 4
        contador_planilha = 1
        for b in range(16, 32):
            time = lista_times[numeros_sorteados[b]]
            plan[f'M{contador_planilha}'] = time[0]
            plan[f'I{contador_planilha}'] = time[1]
            plan[f'I{contador_planilha + 2}'] = time[2]
            contador_planilha += 4
        arquivo.save('ligaroletarussa2021.xlsx')

    def mata_mata_a(self):
        times = self.pegar_times(a=True)
        numeros = self.numeros_aleatorios()
        self.sorteio(lista_times=times, numeros_sorteados=numeros, planilha='MataMataA')
        print('Mata-mata Série A Criado com Sucesso.')

    def mata_mata_b(self):
        times = self.pegar_times(b=True)
        numeros = self.numeros_aleatorios()
        self.sorteio(lista_times=times, numeros_sorteados=numeros, planilha='MataMataB')
        print('Mata-mata Série B Criado com Sucesso.')

    def mata_mata_c(self):
        times = self.pegar_times(c=True)
        numeros = self.numeros_aleatorios()
        self.sorteio(lista_times=times, numeros_sorteados=numeros, planilha='MataMataC')
        print('Mata-mata Série C Criado com Sucesso.')

    def criar_mata_matas(self):
        self.mata_mata_a()
        self.mata_mata_b()
        self.mata_mata_c()   

    def pontuacao_sem_capitao(self):
        api = self.acesso_autenticado()
        arquivo = load_workbook('ligaroletarussa2021.xlsx')
        planilhas = ['MataMataA','MataMataB', 'MataMataC']
        for p in planilhas:
            contador = 1
            planilha = arquivo[p]
            while True:
                id = planilha[f'A{contador}'].value
                if id:
                    time = api.time(id=id, as_json=True)
                    pontos = 0
                    for jogador in time['atletas']:
                        pontos += jogador['pontos_num']
                    planilha[f'F{contador}'] = pontos
                    contador += 4
                else:
                    break
            contador = 1
            while True:
                id = planilha[f'M{contador}'].value
                if id:
                    time = api.time(id=id, as_json=True)
                    pontos = 0
                    for jogador in time['atletas']:
                        pontos += jogador['pontos_num']
                    planilha[f'H{contador}'] = pontos
                    contador += 4
                else:
                    break
            arquivo.save('ligaroletarussa2021.xlsx')
            print(f'Pontuações do {p} atualizadas!')

    def pontuacao_com_capitao(self):
        api = self.acesso_autenticado()
        arquivo = load_workbook('ligaroletarussa2021.xlsx')
        planilhas = ['MataMataA','MataMataB', 'MataMataC']
        for p in planilhas:
            contador = 1
            planilha = arquivo[p]
            while True:
                id = planilha[f'A{contador}'].value
                if id:
                    time = api.time(id=id, as_json=True)
                    pontos = time['pontos']
                    planilha[f'F{contador}'] = pontos
                    contador += 4
                else:
                    break
            contador = 1
            while True:
                id = planilha[f'M{contador}'].value
                if id:
                    time = api.time(id=id, as_json=True)
                    pontos = time['pontos']
                    planilha[f'H{contador}'] = pontos
                    contador += 4
                else:
                    break
            arquivo.save('ligaroletarussa2021.xlsx')
            print(f'Pontuações do {p} atualizadas!')


class MataMataDuplas(Pontuacao):
    def __init__(self):
        super().__init__()
        self.tabela_bd = 'MMDuplas'
    
    def escolher_time(self):
        tela = Exibir()
        while True:
            lista_times = self.pesquisar_time()
            tela.listar_itens(lista_times)
            time = tela.escolher_entre_opcoes(lista_times)
            if time:
                break
        return time

    def cadastrar_duplas(self):
        contador = 1
        tela = Exibir()
        while True:
            tela.exibir_cabecalho(f'Dupla {contador}')
            dupla = str(input('Nome da Dupla: ')).upper().strip()
            tela.exibir_cabecalho(f'{dupla}')
            print('Time 1')
            time1 = self.escolher_time()
            print('Time 2')
            time2 = self.escolher_time()
            con, cursor = self.acessar_banco_de_dados()
            dados = (dupla, time1['id'], time1['nome'], time2['id'], time2['nome'])
            cursor.execute(f'INSERT INTO {self.tabela_bd}(Dupla, ID_Time1, Nome_Time1, ID_Time2, Nome_Time2) VALUES{dados}')
            con.commit()
            print(f"""
            Dupla: {dupla}
            Time 1: {time1['nome']}
            Time 2: {time2['nome']}

            Cadastro Efetuado com Sucesso.
            """)
            while True:
                escolha = str(input('Cadastrar mais Duplas? [S/N] ')).upper().strip()
                if escolha == 'N':
                    exit()
                if escolha == 'S':
                    break
                else:
                    print('Opção inválida. ')
            if contador == 32:
                print('As 32 Duplas já foram cadastradas!')
                break
            else:
                contador += 1

    def numeros_aleatorios(self):
        numeros_sorteados = []
        while True:
            numero = randint(0, 31)
            if numero not in numeros_sorteados:
                numeros_sorteados.append(numero)
            elif len(numeros_sorteados) == 32:
                break
        return numeros_sorteados

    def sorteio(self):
        arquivo = load_workbook('ligaroletarussa2021.xlsx')
        planilha = arquivo['MataMataDuplas']
        numeros = self.numeros_aleatorios()
        con, cursor = self.acessar_banco_de_dados()
        cursor.execute(f'SELECT * FROM {self.tabela_bd}')
        duplas = cursor.fetchall()
        contador_planilha = 1
        for a in range(0, 16):
            time = duplas[numeros[a]]
            planilha[f'F{contador_planilha}'] = time[1]
            planilha[f'A{contador_planilha}'] = time[2]
            planilha[f'B{contador_planilha}'] = time[3]
            planilha[f'A{contador_planilha + 1}'] = time[4]
            planilha[f'B{contador_planilha + 1}'] = time[5]
            contador_planilha += 3
        contador_planilha = 1
        for a in range(16, 32):
            time = duplas[numeros[a]]
            planilha[f'M{contador_planilha}'] = time[1]
            planilha[f'U{contador_planilha}'] = time[2]
            planilha[f'Q{contador_planilha}'] = time[3]
            planilha[f'U{contador_planilha + 1}'] = time[4]
            planilha[f'Q{contador_planilha + 1}'] = time[5]
            contador_planilha += 3
        arquivo.save('ligaroletarussa2021.xlsx')

    def criar_mata_mata_duplas(self):
        self.cadastrar_duplas()
        self.sorteio()
        print('Mata-mata em Duplas criado com sucesso.')
    
    def pontuacao_sem_capitao(self):
        api = self.acesso_autenticado()
        arquivo = load_workbook('ligaroletarussa2021.xlsx')
        planilha = arquivo['MataMataDuplas']
        contador = 1
        while True:
            id1 = planilha[f'A{contador}'].value
            id2 = planilha[f'A{contador + 1}'].value
            if id1:
                time1 = api.time(id=id1, as_json=True)
                time2 = api.time(id=id2, as_json=True)
                pontos = 0
                for j1 in time1['atletas']:
                    pontos += j1['pontos_num']
                for j2 in time2['atletas']:
                    pontos += j2['pontos_num']
                planilha[f'J{contador}'] = pontos
                contador +3
            else:
                break
        contador = 1
        while True:
            id1 = planilha[f'U{contador}'].value
            id2 = planilha[f'U{contador + 1}'].value
            if id1:
                time1 = api.time(id=id1, as_json=True)
                time2 = api.time(id=id2, as_json=True)
                pontos = 0
                for j1 in time1['atletas']:
                    pontos += j1['pontos_num']
                for j2 in time2['atletas']:
                    pontos += j2['pontos_num']
                planilha[f'L{contador}'] = pontos
                contador +3
            else:
                break
        arquivo.save('ligaroletarussa2021.xlsx')
        print(f'Pontuações do MataMataDuplas atualizadas!')

    def pontuacao_com_capitao(self):
        api = self.acesso_autenticado()
        arquivo = load_workbook('ligaroletarussa2021.xlsx')
        planilha = arquivo['MataMataDuplas']
        contador = 1
        while True:
            id1 = planilha[f'A{contador}'].value
            id2 = planilha[f'A{contador + 1}'].value
            pontos = 0
            if id1:
                time1 = api.time(id=id1, as_json=True)
                pontos += time1['pontos']
                time2 = api.time(id=id2, as_json=True)
                pontos += time2['pontos']
                planilha[f'J{contador}'] = pontos
                contador += 3
            else:
                break
        contador = 1
        while True:
            id1 = planilha[f'U{contador}'].value
            id2 = planilha[f'U{contador + 1}'].value
            pontos = 0
            if id1:
                time1 = api.time(id=id1, as_json=True)
                pontos += time1['pontos']
                time2 = api.time(id=id2, as_json=True)
                pontos += time2['pontos']
                planilha[f'L{contador}'] = pontos
                contador += 3
            else:
                break
        arquivo.save('ligaroletarussa2021.xlsx')
        print(f'Pontuações do MataMataDuplas atualizadas!')






